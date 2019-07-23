'''
When a worksheet is created in memory, it contains no cells.
They are created when first accessed.
'''
from openpyxl import Workbook

wb=Workbook()
ws=wb.active
ws['A1']='姓名'
# 可以通过ws['A1']的方式读取单元格内容
c=ws['A1'].value
print(c)

# 还可以通过Worksheet.cell()方法，通过行和列访问
d=ws.cell(row=1,column=2,value=10)
print(d.value)

# 访问多个单元格  方式1  注意：访问单元格返回的是之前定义了的单元格，不是实际的所有单元格
cell_range=ws['A1':'C2']
print(cell_range)

#访问多个单元格,通过行和列  方式2
colc=ws['C']
print(colc)  # 返回C1 、C2单元格
col_range=ws['C:D']
print(col_range)
row10=ws[10]
print(row10)
row_range=ws[5:10]
print(row_range)


# 用Worksheet.iter_rows() 方法  需要制定行 
for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print(cell)
#  用Worksheet.iter_cols() 方法   需要制定列
for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print(cell)

# 可以通过Worksheet.rows 或Worksheet.columns 迭代所有行和列
print(tuple(ws.rows))
print(tuple(ws.columns))

# 访问值，可以用Worksheet.values属性获取全部的单元格的值
for row in ws.values:
    for value in row:
        print(value)
# Worksheet.iter_rows()和Worksheet.iter.cols()方法可以用values_only参数获取单元格的值
for row in ws.iter_rows(min_row=1,max_col=2,max_row=3,values_only=True):
    print(row)
wb.save('单元格操作.xlsx')
