# The same way as writing, you can use the openpyxl.load_workbook()
# to open an existing workbook:
from openpyxl import load_workbook
wb=load_workbook('本地文件.xlsx')
print(wb.sheetnames)
