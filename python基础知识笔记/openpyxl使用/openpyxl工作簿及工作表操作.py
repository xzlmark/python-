from openpyxl import Workbook
import datetime
# 创建工作簿
wb=Workbook()
# 激活wb,产生一个worksheet,默认sheet名字是sheet
ws=wb.active
# 设定sheet的背景颜色,talColor 注意C是大写
ws.sheet_properties.tabColor='1072BA'
# 默认在后面插入sheet
ws1=wb.create_sheet('默认在末尾添加sheet')
# 指定位置插入sheet
ws2=wb.create_sheet('指定位置添加sheet',0)

# 只要给worksheet设定了名字，就可以通过下面这种方式调用
ws3=wb["指定位置添加sheet"]
ws3.sheet_properties.tabColor='FF0000'

# ws.title是在任何时候都可以使用，表示修改sheet的名字
ws.title='第一个sheet'

ws['A1']=42

ws.append([1,2,3])

ws['A2']=datetime.datetime.now()
# 输出所有工作簿的工作表名称
print(wb.sheetnames)
# 可以通过循环输出工作表
for sheet in wb:
    print(sheet.title)

# 在一个工作簿内，可以复制工作表,不可以在工作簿之间复制
target=wb.copy_worksheet(ws)
wb.save('openpyxl工作簿及工作表操作.xlsx')
